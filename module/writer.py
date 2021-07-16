import openpyxl
import os
import pandas as pd

FILE_NAME = 'PRTimes営業リスト.xlsx'
COMPANY_COLUMN = 1
MAILING_COLUMN = 2


class ExcelWriter():
    def __init__(self, path) -> None:
        self.path = os.path.join(path, FILE_NAME)

    def write(self, company_list, mailing_list):
        """書き込み

        Args:
            company_list (list): 会社名リスト
            mailing_list (list): メールアドレスリスト
        Notes:
            ファイルが存在するかどうかで新規書き込みにするか上書きするか判断して書き込みを行う
            内部で振り分けするため呼び出し側は意識しなくていい
        """
        if os.path.isfile(self.path):
            # 追記処理
            self.add_write(company_list, mailing_list)
            # 追記処理の場合は書き込んだ際に重複が存在する可能性があるため
            self.duplicate()
        else:
            # 新規処理
            self.new_write(company_list, mailing_list)

    def add_write(self, company_list, mailing_list):
        """追記書き込み

        Args:
            company_list (list): 会社名リスト
            mailing_list (list): メールアドレスリスト

        Notes:
            上書き処理
        """
        wb = openpyxl.load_workbook(self.path)
        ws = wb.worksheets[0]

        # 最大行
        maxRow = ws.max_row + 1

        # 行を逆ループ
        for i in reversed(range(1, maxRow)):

            # セルの値がNoneじゃなかったら、次の行から書き込み開始
            if ws.cell(row=i, column=COMPANY_COLUMN).value is not None:

                # 配列ループ
                for list_index in range(0, len(company_list)):

                    # リストを書き込み
                    ws.cell(
                        i + 1,
                        COMPANY_COLUMN,
                        value=company_list[list_index])
                    ws.cell(
                        i + 1,
                        MAILING_COLUMN,
                        value=mailing_list[list_index])
                    i = i + 1
                break

        wb.save(self.path)

    def new_write(self, company_list, mailing_list):
        """新規書き込み

        Args:
            company_list (list): 会社名リスト
            mailing_list (list): メールアドレスリスト
        """
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        # カラム名の記載
        ws.cell(1, COMPANY_COLUMN, value='会社名')
        ws.cell(1, MAILING_COLUMN, value='メールアドレス')

        # 配列ループ
        for i in range(0, len(company_list)):

            # リストを書き込み
            ws.cell(
                i + 2,
                COMPANY_COLUMN,
                value=company_list[i])
            ws.cell(
                i + 2,
                MAILING_COLUMN,
                value=mailing_list[i])

        wb.save(self.path)

    def duplicate(self):
        df = pd.read_excel(self.path)
        df.drop_duplicates(inplace=True)
        df.to_excel(self.path, index=False, header=True)
